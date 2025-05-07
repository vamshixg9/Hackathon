from flask import Flask, request, jsonify, session, render_template, redirect, url_for
from authenticate import authenticate
from send_otp import send_otp
from validate import validate_otp  # You'll write this soon
from flask_cors import CORS
import os
from flask_sqlalchemy import SQLAlchemy
from models import db, User, Attendance, OTPLog, Ticket
import openpyxl
from openpyxl.styles import Font
from flask import send_file
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter


app = Flask(__name__)
CORS(app)
app.secret_key = os.urandom(24)
from sqlalchemy import create_engine
# from sqlalchemy.pool import NullPool
from dotenv import load_dotenv
import os

# Load environment variables from .env
load_dotenv()

# Fetch variables
USER = os.getenv("user")
PASSWORD = os.getenv("password")
HOST = os.getenv("host")
PORT = os.getenv("port")
DBNAME = os.getenv("dbname")




basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

with app.app_context():
    db.create_all()

@app.route("/")
def index():
    return render_template("index.html")

@app.route('/login', methods=['POST'])
def login():
    data = request.get_json()
    email = data.get('email')
    password = data.get('password')

    if not email.endswith('@morabu.com'):
        return jsonify({'success': False, 'message': 'Invalid email domain'}), 400

    if not authenticate(email, password):
        return jsonify({'success': False, 'message': 'Invalid credentials'}), 401
    
     # Check if user exists
    user = User.query.filter_by(email=email).first()
    if not user:
        # Insert user with dummy hash (replace later with real hash)
        user = User(email=email, password_hash="placeholder_hashed_pw", role="employee", name="Unknown")
        db.session.add(user)
        db.session.commit()

    # Send OTP
    otp = send_otp('diksha@morabu.com', 'uzqgy48b', email)
    session['otp'] = otp
    session['email'] = email

    # Save OTP in OTPLog (optional but useful)
    otp_entry = OTPLog(user_id=user.id, otp_code=otp)
    db.session.add(otp_entry)
    db.session.commit()

    return jsonify({'success': True, 'message': 'OTP sent to your email'}), 200


@app.route("/verify_otp", methods=["POST"])
def verify_otp():
    data = request.json
    entered_otp = data["otp"]
    saved_otp = session.get("otp")
    user_email = session.get("email")

    if entered_otp != saved_otp:
        return jsonify(success=False, message="Incorrect OTP.")

    user = User.query.filter_by(email=user_email).first()
    if not user:
        return jsonify(success=False, message="User not found.")

    session['role'] = user.role  # Store user role for session use

    # Decide redirect based on role
    if user.role == "admin":
        redirect_url = url_for("admin_home")
    else:
        redirect_url = url_for("home")

    return jsonify(success=True, redirect_url=redirect_url)

@app.route("/admin_home")
def admin_home():
    user_email = session.get('email')
    user = User.query.filter_by(email=user_email).first()

    if not user or user.role != "admin":
        return redirect(url_for('index'))

    return render_template("admin_home.html")

@app.route("/user")
def user():
    return render_template("user.html")

@app.route("/tickets")
def tickets():
    return render_template("tickets.html")

@app.route("/list")
def list_items():
    return render_template("list.html")

@app.route("/admin_user")
def admin_user():
    return render_template("admin_user.html")

@app.route("/admin_tickets")
def admin_tickets():
    return render_template("admin_tickets.html")

@app.route("/admin_offices")
def admin_offices():
    return render_template("admin_offices.html")

@app.route("/admin_employees")
def admin_employees():
    return render_template("admin_employees.html")

@app.route("/admin_admin")
def admin_admin():
    return render_template("admin_admin.html")



@app.route("/home")
def home():  # ← WRONG function name or return
    user_email = session.get('email')  # Example of retrieving the logged-in user's email
    user = User.query.filter_by(email=user_email).first()
    print(user)
    if user:
        # Get the attendance records of the user, maybe limit it to a certain date range or show all
        logs = Attendance.query.filter_by(user_id=user.id).all()
        return render_template("home.html", logs=logs)  # Passing the logs to the template
    return redirect(url_for('index'))

@app.route('/timelogs')
def timelogs():
    user_email = session.get('email')  # Retrieve logged-in user's email
    user = User.query.filter_by(email=user_email).first()
    
    if user:
        # Get attendance records specific to the user
        logs = Attendance.query.filter_by(user_id=user.id).all()
        return render_template("timelogs.html", logs=logs)  # Pass logs to template
    return redirect(url_for('index'))

from datetime import timedelta

def calculate_total_hours(checkin, checkout):
    if checkin and checkout:
        total = checkout - checkin
        hours, remainder = divmod(total.total_seconds(), 3600)
        minutes = remainder // 60
        return f"{int(hours):02}:{int(minutes):02}"
    return "Active" if checkin and not checkout else ""

@app.route('/api/attendance', methods=['GET'])
def get_attendance():
    attendance_records = Attendance.query.all()

    attendance_data = [
        {
            "Date": record.date.strftime("%Y-%m-%d"),
            "ClockIn": record.checkin_time.strftime("%H:%M") if record.checkin_time else "",
            "ClockOut": record.checkout_time.strftime("%H:%M") if record.checkout_time else "",
            "TotalHrs": calculate_total_hours(record.checkin_time, record.checkout_time),
            "Location": record.location or ""
        }
        for record in attendance_records
    ]

    return jsonify({"attendance": attendance_data})


@app.route('/download/excel')
def download_excel():
    records = Attendance.query.all()

    # Assuming you want the first user in the attendance records for now
    # You can modify this to get the specific user by email or other identifiers
    user_email = session.get('email')
    user = User.query.filter_by(email=user_email).first()


    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Attendance Records"

    # Add Title Row
    title = "MORABU HANSHIN : TIME SHEET"
    sheet.merge_cells('A1:E1')
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = title
    title_cell.font = Font(size=14, bold=True)

    # Add User's Name on Top
    user_name = user.name if user else "Unknown User"
    sheet.merge_cells('A2:E2')
    user_name_cell = sheet.cell(row=2, column=1)
    user_name_cell.value = f"User: {user_name}"
    user_name_cell.font = Font(size=12, bold=True)

    # Add Headers
    headers = ["Date", "Clock In", "Clock Out", "Location", "Total Hours"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)

    # Add Data Rows
    for row_idx, record in enumerate(records, start=4):
        sheet.cell(row=row_idx, column=1, value=record.date.strftime("%Y-%m-%d"))
        sheet.cell(row=row_idx, column=2, value=record.checkin_time.strftime("%H:%M") if record.checkin_time else "")
        sheet.cell(row=row_idx, column=3, value=record.checkout_time.strftime("%H:%M") if record.checkout_time else "")
        sheet.cell(row=row_idx, column=4, value=record.location or "")
        sheet.cell(row=row_idx, column=5, value=calculate_total_hours(record.checkin_time, record.checkout_time))

    # Save to memory stream
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(output,
                     download_name="timesheet_モラブ阪神工業株式会社.xlsx",
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas


from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import fonts

@app.route('/download/pdf')
def download_pdf():
    records = Attendance.query.all()
    filename = "timesheet_モラブ阪神工業株式会社.pdf"
    pdf_file = BytesIO()

    c = canvas.Canvas(pdf_file, pagesize=letter)

    # Title and User's Name
    c.drawString(200, 750, "MORABU HANSHIN : TIME SHEET")

    # Assuming you want the first user in the attendance records for now
    user_email = session.get('email')
    user = User.query.filter_by(email=user_email).first()


    # User's Name
    user_name = user.name if user else "Unknown User"
    c.drawString(50, 730, f"User: {user_name}")

    # Adjust the y_position for the headers
    y_position = 710

    # Table headers
    c.drawString(50, y_position, "Date")
    c.drawString(150, y_position, "Clock In")
    c.drawString(250, y_position, "Clock Out")
    c.drawString(350, y_position, "Location")
    c.drawString(450, y_position, "Total Hours")

    y_position -= 20  # Move down after headers to avoid overlap

    # Loop through records and add them to the table
    for record in records:
        # Format times as HH:MM for cleaner presentation
        checkin_time = record.checkin_time.strftime('%H:%M') if record.checkin_time else ''
        checkout_time = record.checkout_time.strftime('%H:%M') if record.checkout_time else ''

        # Draw each piece of data in a separate column with sufficient spacing
        c.drawString(50, y_position, str(record.date))  # Date
        c.drawString(150, y_position, checkin_time)  # Clock In (formatted time)
        c.drawString(250, y_position, checkout_time)  # Clock Out (formatted time)
        c.drawString(350, y_position, record.location if record.location else '')  # Location
        c.drawString(450, y_position, str(record.total_hours))  # Total hours

        y_position -= 20  # Move down for the next row

        # If the table is getting too long, add a page break
        if y_position < 100:
            c.showPage()  # Start a new page
            c.setFont("NotoSansCJK-Regular", 12)  # Reset font for new page
            y_position = 750  # Reset y_position for new page
            # Re-add headers on new page
            c.drawString(50, 730, "Date")
            c.drawString(150, 730, "Clock In")
            c.drawString(250, 730, "Clock Out")
            c.drawString(350, 730, "Location")
            c.drawString(450, 730, "Total Hours")

    c.save()

    pdf_file.seek(0)
    return send_file(pdf_file, as_attachment=True, download_name=filename, mimetype='application/pdf')

@app.route('/api/tickets', methods=['GET', 'POST'])
def handle_tickets():
    user_email = session.get('email')
    user = User.query.filter_by(email=user_email).first()

    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    if request.method == 'POST':
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Missing JSON data'}), 400

        title = data.get('title')
        description = data.get('description')

        if not title or not description:
            return jsonify({'error': 'Missing required fields'}), 400

        new_ticket = Ticket(title=title, description=description, user_id=user.id)
        db.session.add(new_ticket)
        db.session.commit()

        return jsonify({
            'success': True,
            'ticket': {
                'id': new_ticket.id,
                'user_id': new_ticket.user_id,
                'title': new_ticket.title,
                'description': new_ticket.description,
                'status': new_ticket.status,
                'created_at': new_ticket.created_at.strftime("%Y-%m-%d %H:%M:%S")
            }
        }), 201

    else:  # GET request
        tickets = Ticket.query.filter_by(user_id=user.id).all()
        return jsonify([{
            'id': t.id,
            'title': t.title,
            'description': t.description,
            'status': t.status,
            'created_at': t.created_at.strftime("%Y-%m-%d %H:%M:%S")
        } for t in tickets])


@app.route('/api/tickets/<int:ticket_id>', methods=['DELETE'])
def delete_ticket(ticket_id):
    user_email = session.get('email')
    user = User.query.filter_by(email=user_email).first()

    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    ticket = Ticket.query.get(ticket_id)

    if not ticket or ticket.user_id != user.id:
        return jsonify({'error': 'Ticket not found or unauthorized'}), 404

    db.session.delete(ticket)
    db.session.commit()

    return jsonify({'success': True}), 200


from flask import Blueprint, request, jsonify, session
from models import db, ToDo, User

todo_bp = Blueprint('todos', __name__)

@todo_bp.route('/api/todos', methods=['GET'])
def get_todos():
    user_email = session.get('email')
    user = User.query.filter_by(email=user_email).first()

    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    todos = ToDo.query.filter_by(user_id=user.id).order_by(ToDo.created_at.desc()).all()
    return jsonify({'todos': [t.to_dict() for t in todos]})

@todo_bp.route('/api/todos', methods=['POST'])
def add_todo():
    user_email = session.get('email')
    print(f"User email from session: {user_email}")
    user = User.query.filter_by(email=user_email).first()

    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    content = request.get_json().get('content', '').strip()
    if not content:
        return jsonify({'error': 'Content required'}), 400

    todo = ToDo(user_id=user.id, content=content)
    db.session.add(todo)
    db.session.commit()
    return jsonify({'todo': todo.to_dict()})

@todo_bp.route('/api/todos/<int:todo_id>/toggle', methods=['POST'])
def toggle_todo(todo_id):
    user_email = session.get('email')
    user = User.query.filter_by(email=user_email).first()

    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    todo = ToDo.query.filter_by(id=todo_id, user_id=user.id).first_or_404()
    todo.is_done = not todo.is_done
    db.session.commit()
    return jsonify({'todo': todo.to_dict()})

app.register_blueprint(todo_bp)

# @app.route('/api/tickets', methods=['POST'])
# def create_ticket():
#     data = request.get_json()

#     user_id = data.get('user_id')  # this must be provided by frontend
#     title = data.get('title')
#     description = data.get('description')

#     if not user_id or not title or not description:
#         return jsonify({'error': 'Missing required fields'}), 400

#     new_ticket = Ticket(
#         user_id=user_id,
#         title=title,
#         description=description,
#         status='open',
#         created_at=datetime.utcnow()
#     )

#     db.session.add(new_ticket)
#     db.session.commit()

#     return jsonify({'message': 'Ticket created successfully', 'ticket_id': new_ticket.id}), 201

# @app.route('/api/tickets', methods=['GET'])
# def get_tickets():
#     tickets = Ticket.query.order_by(Ticket.created_at.desc()).all()

#     result = []
#     for ticket in tickets:
#         result.append({
#             'id': ticket.id,
#             'user_id': ticket.user_id,
#             'title': ticket.title,
#             'description': ticket.description,
#             'status': ticket.status,
#             'created_at': ticket.created_at.strftime('%Y-%m-%d %H:%M:%S'),
#             'updated_at': ticket.updated_at.strftime('%Y-%m-%d %H:%M:%S') if ticket.updated_at else None,
#             'assigned_to': ticket.assigned_to
#         })

#     return jsonify(result), 200

@app.route('/api/right-panel')
def get_right_panel():
    user = User.query.filter_by(email=session.get("email")).first()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    return jsonify({
        "user": {
            "name": user.name,
            "department": user.department,
            "employee_id": user.id,
            "profile_pic": url_for('static', filename='profile-1.jpg')  # fallback image
        }
    })
@app.route('/api/right-panel-todos')
def right_panel_todos():
    user_email = session.get('email')
    if not user_email:
        return jsonify({"error": "Unauthorized"}), 401

    user = User.query.filter_by(email=user_email).first()
    if not user:
        return jsonify({"error": "User not found"}), 404

    # Fetch the 2 oldest uncompleted todos
    todos = ToDo.query.filter_by(user_id=user.id, is_done=False).order_by(ToDo.created_at.asc()).limit(2).all()

    return jsonify({
        "todos": [{"id": t.id, "content": t.content} for t in todos]
    })


@app.route('/api/ticket-count')
def ticket_count():
    user_email = session.get('email')
    print("[DEBUG] Session email:", user_email)

    user = User.query.filter_by(email=user_email).first()
    print("[DEBUG] User:", user)

    if not user:
        return jsonify({'count': 0})

    count = Ticket.query.filter_by(user_id=user.id).count()
    print("[DEBUG] Ticket count:", count)

    return jsonify({'count': count})





import uuid
from datetime import date
from flask import session, jsonify
from datetime import datetime
from models import db, Attendance, User


@app.route('/api/admin/scan', methods=['POST'])
def admin_scan():
    print("waiting for QR")
    data = request.get_json()
    token = data.get('token')
    print(f"Scanned token: {token}")



    if not token:
        return jsonify({'success': False, 'message': 'Missing token'}), 400

    # Find the matching attendance record
    attendance = Attendance.query.filter_by(qr_token=token).first()

    if not attendance:
        return jsonify({'success': False, 'message': 'Invalid or expired token'}), 404

    try:
        if attendance.checkin_time is None:
            # Check-in flow
            attendance.checkin_time = datetime.utcnow()
            attendance.qr_token = None
            message = "Check-in successful"

        elif attendance.checkout_time is None:
            # Check-out flow
            attendance.checkout_time = datetime.utcnow()
            attendance.qr_token = None
            message = "Check-out successful"

        else:
            # Already checked in and out
            return jsonify({'success': False, 'message': 'Already checked out'}), 400

        db.session.commit()
        return jsonify({'success': True, 'message': message})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Server error: {str(e)}'}), 500
    

@app.route('/api/qr_status/<token>')
def qr_status(token):
    record = Attendance.query.filter_by(qr_token=token).first()

    if not record:
        # Token is no longer in DB: likely used and cleared
        return jsonify({'used': True})

    # If either time is populated, token was acted on
    if record.checkin_time or record.checkout_time:
        return jsonify({'used': True})

    return jsonify({'used': False})


@app.route('/api/generate_qr', methods=['POST'])
def generate_qr():
    user_email = session.get('email')
    user = User.query.filter_by(email=user_email).first()
    
    if not user:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401

    today = date.today()

    # Find latest attendance entry
    latest_attendance = (
        Attendance.query.filter_by(user_id=user.id)
        .order_by(Attendance.id.desc())
        .first()
    )

    # ✅ Create new only if last one is missing or already checked out
    if not latest_attendance or latest_attendance.checkout_time:
        new_attendance = Attendance(
            user_id=user.id,
            date=today,
            qr_token= str(uuid.uuid4())
        )
        db.session.add(new_attendance)
        db.session.commit()
        token = new_attendance.qr_token
    else:
        # ❌ Reuse existing (not yet checked out)
        latest_attendance.qr_token = str(uuid.uuid4())
        db.session.commit()
        token = latest_attendance.qr_token

    return jsonify({'success': True, 'token': token})


@app.route('/api/expire_qr', methods=['POST'])
def expire_qr():
    user_email = session.get('email')
    if not user_email:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401

    user = User.query.filter_by(email=user_email).first()
    if not user:
        return jsonify({'success': False, 'message': 'User not found'}), 404

    today = date.today()
    attendance = Attendance.query.filter_by(user_id=user.id, date=today).first()
    if attendance:
        attendance.qr_token = None
        db.session.commit()

    return jsonify({'success': True})


from flask import jsonify, session
from datetime import date
from models import Attendance, User

@app.route('/api/user/attendance-status')
def attendance_status():
    user_email = session.get('email')
    if not user_email:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401

    user = User.query.filter_by(email=user_email).first()
    if not user:
        return jsonify({'success': False, 'message': 'User not found'}), 404

    today = date.today()
    record = Attendance.query.filter_by(user_id=user.id, date=today).first()

    if record:
        return jsonify({
            'success': True,
            'checked_in': record.checkin_time is not None,
            'checkin_time': record.checkin_time.isoformat() if record.checkin_time else None,
            'checkout_time': record.checkout_time.isoformat() if record.checkout_time else None,
            'total_hours': record.total_hours
        })

    return jsonify({
        'success': True,
        'checked_in': False,
        'checkin_time': None,
        'checkout_time': None,
        'total_hours': 0.0
    })



@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))


if __name__ == '__main__':
    app.run(debug=True)


