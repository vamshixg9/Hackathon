from flask import Flask, request, jsonify, session, render_template, redirect, url_for
from send_otp import send_otp
from validate import validate_otp  # You'll write this soon
from flask_cors import CORS
import os
from flask_sqlalchemy import SQLAlchemy
from models import db, User, Attendance, OTPLog, Ticket
import openpyxl
from openpyxl.styles import Font
from flask import send_file
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter


app = Flask(__name__)
CORS(app)
app.secret_key = os.urandom(24)
from sqlalchemy import create_engine
# from sqlalchemy.pool import NullPool
from dotenv import load_dotenv
import os

# Load environment variables from .env
load_dotenv()

# Fetch variables
USER = os.getenv("user")
PASSWORD = os.getenv("password")
HOST = os.getenv("host")
PORT = os.getenv("port")
DBNAME = os.getenv("dbname")




basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

with app.app_context():
    db.create_all()

@app.route("/")
def index():
    return render_template("index.html")


import smtplib
import random
from email.mime.text import MIMEText

def authenticate(email_input: str, password_input: str) -> bool:
    try:
        with smtplib.SMTP("morabu.com", 587, timeout=5) as server:
            server.starttls()
            server.login(email_input, password_input)
            return True
    except smtplib.SMTPAuthenticationError as e:
        print("[!] SMTP Authentication failed.")
        print(f"Code: {e.smtp_code}, Message: {e.smtp_error.decode()}")
        return False
    except Exception as e:
        print(f"[!] Other error: {e}")
        return False

@app.route('/login', methods=['POST'])
def login():
    data = request.get_json()
    email = data.get('email')
    password = data.get('password')

    if not email:
        return jsonify({'success': False, 'message': 'Email is required'}), 400

    is_guest = password == "guest"

    # Block non-morabu domains for regular users only
    if not is_guest and not email.endswith('@morabu.com'):
        return jsonify({'success': False, 'message': 'Invalid email domain'}), 400

    # Authenticate only for non-guest users
    if not is_guest and not authenticate(email, password):
        return jsonify({'success': False, 'message': 'Invalid credentials'}), 401

    # Check if user exists in DB
    user = User.query.filter_by(email=email).first()
    if not user:
        office_id = None

        if is_guest:
            from models import Office
            latest_office = Office.query.order_by(Office.id.desc()).first()
            if latest_office:
                office_id = latest_office.id

        user = User(
            email=email,
            password_hash="guest" if is_guest else "placeholder_hashed_pw",
            role="employee",
            name="Guest" if is_guest else "Unknown",
            office_id=office_id  # ✅ Assign the latest office
        )
        db.session.add(user)
        db.session.commit()


    # Generate and send OTP
    otp = send_otp('diksha@morabu.com', 'uzqgy48b', email)
    session['otp'] = otp
    session['email'] = email

    # Log OTP for auditing
    otp_entry = OTPLog(user_id=user.id, otp_code=otp)
    db.session.add(otp_entry)
    db.session.commit()

    return jsonify({'success': True, 'message': 'OTP sent to your email'}), 200



@app.route("/verify_otp", methods=["POST"])
def verify_otp():
    data = request.json
    entered_otp = data["otp"]
    saved_otp = session.get("otp")
    user_email = session.get("email")

    if entered_otp != saved_otp:
        return jsonify(success=False, message="Incorrect OTP.")

    user = User.query.filter_by(email=user_email).first()
    if not user:
        return jsonify(success=False, message="User not found.")

    session['role'] = user.role  # Store user role for session use

    # Decide redirect based on role
    if user.role == "admin":
        redirect_url = url_for("admin_home")
    else:
        redirect_url = url_for("home")

    return jsonify(success=True, redirect_url=redirect_url)

@app.route("/admin_home")
def admin_home():
    user_email = session.get('email')
    user = User.query.filter_by(email=user_email).first()

    if not user or user.role != "admin":
        return redirect(url_for('index'))

    return render_template("admin_home.html")

@app.route("/user")
def user():
    user_email = session.get('email')
    user = User.query.filter_by(email=user_email).first()
    if not user:
        return redirect(url_for('index'))

    return render_template("user.html", user=user)



import os
from werkzeug.utils import secure_filename

UPLOAD_FOLDER = os.path.join('static', 'uploads')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/upload_profile', methods=['POST'])
def upload_profile():
    user_email = session.get('email')
    user = User.query.filter_by(email=user_email).first()

    if not user:
        return redirect(url_for('index'))

    user.employee_id = request.form.get('employee_id', user.employee_id)
    user.name = request.form.get('name', user.name)
    user.department = request.form.get('department', user.department)
    user.role = request.form.get('role', user.role)

    if 'photo' in request.files:
        file = request.files['photo']
        if file and file.filename:
            filename = secure_filename(file.filename)
            upload_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(upload_path)
            user.profile_pic = f"/static/uploads/{filename}"

    db.session.commit()
    return redirect(url_for('user'))


@app.route('/upload_Adminprofile', methods=['POST'])
def upload_Adminprofile():
    user_email = session.get('email')
    admin = User.query.filter_by(email=user_email).first()

    if not user:
        return redirect(url_for('index'))
    admin.employee_id = request.form.get('employee_id', admin.employee_id)
    admin.name = request.form.get('name', admin.name)
    admin.department = request.form.get('department', admin.department)
    admin.role = request.form.get('role', admin.role)

    if 'photo' in request.files:
        file = request.files['photo']
        if file and file.filename:
            filename = secure_filename(file.filename)
            upload_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(upload_path)
            user.profile_pic = f"/static/uploads/{filename}"

    db.session.commit()
    return redirect(url_for('user'))


@app.route("/tickets")
def tickets():
    return render_template("tickets.html")

@app.route("/list")
def list_items():
    return render_template("list.html")


@app.route("/admin_tickets")
def admin_tickets():
    return render_template("admin_tickets.html")




@app.route("/admin_user")
def admin_user():
    user_email = session.get('email')
    admin = User.query.filter_by(email=user_email, role="admin").first()

    if not admin:
        return redirect(url_for('index'))

    return render_template("admin_user.html", admin=admin)


@app.route("/admin_admin", methods=["GET", "POST"])
def admin_admin():
    user_email = session.get('email')
    current_user = User.query.filter_by(email=user_email).first()

    if not current_user or current_user.role != "admin":
        return redirect(url_for("index"))

    if request.method == "POST":
        selected_id = request.form.get("employee_id")
        action = request.form.get("action")

        selected_user = User.query.get(selected_id)
        if selected_user:
            if action == "make":
                selected_user.role = "admin"
            elif action == "remove":
                selected_user.role = "employee"
            db.session.commit()
            return redirect(url_for("admin_admin"))

    # Show all users, even admins (so we can demote them)
    employees = User.query.all()

    return render_template("admin_admin.html", employees=employees)


@app.route("/home")
def home():  # ← WRONG function name or return
    user_email = session.get('email')  # Example of retrieving the logged-in user's email
    user = User.query.filter_by(email=user_email).first()
    print(user)
    if user:
        # Get the attendance records of the user, maybe limit it to a certain date range or show all
        logs = Attendance.query.filter_by(user_id=user.id).all()
        return render_template("home.html", logs=logs)  # Passing the logs to the template
    return redirect(url_for('index'))

@app.route('/timelogs')
def timelogs():
    user_email = session.get('email')  # Retrieve logged-in user's email
    user = User.query.filter_by(email=user_email).first()
    
    if user:
        # Get attendance records specific to the user
        logs = Attendance.query.filter_by(user_id=user.id).all()
        return render_template("timelogs.html", logs=logs)  # Pass logs to template
    return redirect(url_for('index'))

from datetime import timedelta

from pytz import timezone
jst = timezone("Asia/Tokyo")

def to_jst(dt):
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=pytz.utc)
    return dt.astimezone(jst)

def calculate_total_hours(checkin, checkout):
    if checkin and checkout:
        total = checkout - checkin
        hours, remainder = divmod(total.total_seconds(), 3600)
        minutes = remainder // 60
        return f"{int(hours):02}:{int(minutes):02}"
    return "Active" if checkin and not checkout else ""


@app.route('/api/attendance', methods=['GET'])
def get_attendance():
    attendance_records = Attendance.query.all()

    attendance_data = [
        {
            "Date": record.date.strftime("%Y-%m-%d"),
            "ClockIn": record.checkin_time.strftime("%H:%M") if record.checkin_time else "",
            "ClockOut": record.checkout_time.strftime("%H:%M") if record.checkout_time else "",
            "TotalHrs": calculate_total_hours(to_jst(record.checkin_time), to_jst(record.checkout_time)),
            "Location": record.location or ""
        }
        for record in attendance_records
    ]

    return jsonify({"attendance": attendance_data})


from pytz import timezone
from io import BytesIO
from flask import send_file

@app.route('/download/excel')
def download_excel():
    jst = timezone('Asia/Tokyo')

    def format_jst(dt):
        if dt:
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=pytz.utc)
            return dt.astimezone(jst).strftime('%H:%M')
        return ""

    user_email = session.get('email')
    user = User.query.filter_by(email=user_email).first()
    records = Attendance.query.filter_by(user_id=user.id).all() 

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Attendance Records"

    # Title
    sheet.merge_cells('A1:E1')
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = "MORABU HANSHIN : TIME SHEET"
    title_cell.font = Font(size=14, bold=True)

    # User Info
    user_name = user.name if user else "Unknown User"
    sheet.merge_cells('A2:E2')
    user_name_cell = sheet.cell(row=2, column=1)
    user_name_cell.value = f"User: {user_name}"
    user_name_cell.font = Font(size=12, bold=True)

    # Headers
    headers = ["Date", "Clock In", "Clock Out", "Location", "Total Hours"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)

    # Rows
    for row_idx, record in enumerate(records, start=4):
        sheet.cell(row=row_idx, column=1, value=record.date.strftime("%Y-%m-%d"))
        sheet.cell(row=row_idx, column=2, value=format_jst(record.checkin_time))
        sheet.cell(row=row_idx, column=3, value=format_jst(record.checkout_time))
        office_name = record.office.name if record.office else "Unknown"
        sheet.cell(row=row_idx, column=4, value=office_name)

        sheet.cell(row=row_idx, column=5, value=calculate_total_hours(record.checkin_time, record.checkout_time))

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="timesheet_モラブ阪神工業株式会社.xlsx",
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )



from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas


from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import fonts
from pytz import timezone
from io import BytesIO
from flask import send_file
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

@app.route('/download/pdf')
def download_pdf():
    jst = timezone('Asia/Tokyo')

    def format_jst(dt):
        if dt:
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=pytz.utc)
            return dt.astimezone(jst).strftime('%H:%M')
        return ""

    user_email = session.get('email')
    user = User.query.filter_by(email=user_email).first()
    records = Attendance.query.filter_by(user_id=user.id).all()
    filename = "timesheet_モラブ阪神工業株式会社.pdf"
    pdf_file = BytesIO()
    c = canvas.Canvas(pdf_file, pagesize=letter)

    # Title and User Info
    c.drawString(200, 750, "MORABU HANSHIN : TIME SHEET")
    user_email = session.get('email')
    user = User.query.filter_by(email=user_email).first()
    user_name = user.name if user else "Unknown User"
    c.drawString(50, 730, f"User: {user_name}")

    y_position = 710

    # Headers
    c.drawString(50, y_position, "Date")
    c.drawString(150, y_position, "Clock In")
    c.drawString(250, y_position, "Clock Out")
    
    c.drawString(350, y_position, "Location")
    c.drawString(450, y_position, "Total Hours")
    y_position -= 20

    for record in records:
        c.drawString(50, y_position, str(record.date))
        c.drawString(150, y_position, format_jst(record.checkin_time))
        c.drawString(250, y_position, format_jst(record.checkout_time))
        office_name = record.office.name if record.office else "Unknown"
        c.drawString(350, y_position, office_name)
        c.drawString(450, y_position, str(record.total_hours))

        y_position -= 20

        if y_position < 100:
            c.showPage()
            y_position = 750
            c.drawString(50, 730, "Date")
            c.drawString(150, 730, "Clock In")
            c.drawString(250, 730, "Clock Out")
            c.drawString(350, 730, "Location")
            c.drawString(450, 730, "Total Hours")

    c.save()
    pdf_file.seek(0)

    return send_file(
        pdf_file,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )


@app.route('/api/tickets', methods=['GET', 'POST'])
def handle_tickets():
    user_email = session.get('email')
    user = User.query.filter_by(email=user_email).first()

    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    if request.method == 'POST':
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Missing JSON data'}), 400

        title = data.get('title')
        description = data.get('description')

        if not title or not description:
            return jsonify({'error': 'Missing required fields'}), 400

        new_ticket = Ticket(title=title, description=description, user_id=user.id)
        db.session.add(new_ticket)
        db.session.commit()

        return jsonify({
            'success': True,
            'ticket': {
                'id': new_ticket.id,
                'user_id': new_ticket.user_id,
                'title': new_ticket.title,
                'description': new_ticket.description,
                'status': new_ticket.status,
                'created_at': new_ticket.created_at.strftime("%Y-%m-%d %H:%M:%S")
            }
        }), 201

    else:  # GET request
        tickets = Ticket.query.filter_by(user_id=user.id).all()
        return jsonify([{
            'id': t.id,
            'title': t.title,
            'description': t.description,
            'status': t.status,
            'created_at': t.created_at.strftime("%Y-%m-%d %H:%M:%S")
        } for t in tickets])


@app.route('/api/tickets/<int:ticket_id>', methods=['DELETE'])
def delete_ticket(ticket_id):
    user_email = session.get('email')
    user = User.query.filter_by(email=user_email).first()

    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    ticket = Ticket.query.get(ticket_id)

    if not ticket or ticket.user_id != user.id:
        return jsonify({'error': 'Ticket not found or unauthorized'}), 404

    db.session.delete(ticket)
    db.session.commit()

    return jsonify({'success': True}), 200


from flask import Blueprint, request, jsonify, session
from models import db, ToDo, User

todo_bp = Blueprint('todos', __name__)

@todo_bp.route('/api/todos', methods=['GET'])
def get_todos():
    user_email = session.get('email')
    user = User.query.filter_by(email=user_email).first()

    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    todos = ToDo.query.filter_by(user_id=user.id).order_by(ToDo.created_at.desc()).all()
    return jsonify({'todos': [t.to_dict() for t in todos]})

@todo_bp.route('/api/todos', methods=['POST'])
def add_todo():
    user_email = session.get('email')
    print(f"User email from session: {user_email}")
    user = User.query.filter_by(email=user_email).first()

    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    content = request.get_json().get('content', '').strip()
    if not content:
        return jsonify({'error': 'Content required'}), 400

    todo = ToDo(user_id=user.id, content=content)
    db.session.add(todo)
    db.session.commit()
    return jsonify({'todo': todo.to_dict()})

@todo_bp.route('/api/todos/<int:todo_id>/toggle', methods=['POST'])
def toggle_todo(todo_id):
    user_email = session.get('email')
    user = User.query.filter_by(email=user_email).first()

    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    todo = ToDo.query.filter_by(id=todo_id, user_id=user.id).first_or_404()
    todo.is_done = not todo.is_done
    db.session.commit()
    return jsonify({'todo': todo.to_dict()})

app.register_blueprint(todo_bp)

# @app.route('/api/tickets', methods=['POST'])
# def create_ticket():
#     data = request.get_json()

#     user_id = data.get('user_id')  # this must be provided by frontend
#     title = data.get('title')
#     description = data.get('description')

#     if not user_id or not title or not description:
#         return jsonify({'error': 'Missing required fields'}), 400

#     new_ticket = Ticket(
#         user_id=user_id,
#         title=title,
#         description=description,
#         status='open',
#         created_at=datetime.utcnow()
#     )

#     db.session.add(new_ticket)
#     db.session.commit()

#     return jsonify({'message': 'Ticket created successfully', 'ticket_id': new_ticket.id}), 201

# @app.route('/api/tickets', methods=['GET'])
# def get_tickets():
#     tickets = Ticket.query.order_by(Ticket.created_at.desc()).all()

#     result = []
#     for ticket in tickets:
#         result.append({
#             'id': ticket.id,
#             'user_id': ticket.user_id,
#             'title': ticket.title,
#             'description': ticket.description,
#             'status': ticket.status,
#             'created_at': ticket.created_at.strftime('%Y-%m-%d %H:%M:%S'),
#             'updated_at': ticket.updated_at.strftime('%Y-%m-%d %H:%M:%S') if ticket.updated_at else None,
#             'assigned_to': ticket.assigned_to
#         })

#     return jsonify(result), 200

@app.route('/api/right-panel')
def get_right_panel():
    user = User.query.filter_by(email=session.get("email")).first()
    if not user:
        return jsonify({"error": "Unauthorized"}), 401

    return jsonify({
        "user": {
            "name": user.name,
            "department": user.department,
            "employee_id": user.employee_id,
            "profile_pic": url_for('static', filename='profile-1.jpg')  # fallback image
        }
    })
@app.route('/api/right-panel-todos')
def right_panel_todos():
    user_email = session.get('email')
    if not user_email:
        return jsonify({"error": "Unauthorized"}), 401

    user = User.query.filter_by(email=user_email).first()
    if not user:
        return jsonify({"error": "User not found"}), 404

    # Fetch the 2 oldest uncompleted todos
    todos = ToDo.query.filter_by(user_id=user.id, is_done=False).order_by(ToDo.created_at.asc()).limit(2).all()

    return jsonify({
        "todos": [{"id": t.id, "content": t.content} for t in todos]
    })

@app.route('/api/user/attendance-logs')
def user_attendance_logs():
    from pytz import timezone
    jst = timezone('Asia/Tokyo')

    def to_jst(dt):
        if dt is None:
            return None
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=pytz.utc)
        return dt.astimezone(jst).isoformat()

    user_email = session.get('email')
    if not user_email:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401

    user = User.query.filter_by(email=user_email).first()
    if not user:
        return jsonify({'success': False, 'message': 'User not found'}), 404

    logs = (
        Attendance.query
        .filter_by(user_id=user.id)
        .order_by(Attendance.date.desc())
        .limit(30)
        .all()
    )

    return jsonify({
        'success': True,
        'logs': [
            {
                'date': att.date.isoformat(),
                'checkin_time': to_jst(att.checkin_time),
                'checkout_time': to_jst(att.checkout_time),
                'total_hours': att.total_hours
            }
            for att in logs
        ]
    })




@app.route('/api/ticket-count')
def ticket_count():
    user_email = session.get('email')
    print("[DEBUG] Session email:", user_email)

    user = User.query.filter_by(email=user_email).first()
    print("[DEBUG] User:", user)

    if not user:
        return jsonify({'count': 0})

    count = Ticket.query.filter_by(user_id=user.id).count()
    print("[DEBUG] Ticket count:", count)

    return jsonify({'count': count})





import uuid
from datetime import date
from flask import session, jsonify
from datetime import datetime
from models import db, Attendance, User


@app.route('/api/admin/scan', methods=['POST'])
def admin_scan():
    print("waiting for QR")
    data = request.get_json()
    token = data.get('token')
    print(f"Scanned token: {token}")



    if not token:
        return jsonify({'success': False, 'message': 'Missing token'}), 400

    # Find the matching attendance record
    attendance = Attendance.query.filter_by(qr_token=token).first()

    if not attendance:
        return jsonify({'success': False, 'message': 'Invalid or expired token'}), 404

    try:
        from datetime import datetime
        from zoneinfo import ZoneInfo  # Available in Python 3.9+

        tokyo_tz = ZoneInfo("Asia/Tokyo")

        if attendance.checkin_time is None:
            # Check-in flow
            attendance.checkin_time = datetime.now(tokyo_tz)
            attendance.qr_token = None
            message = "Check-in successful"
            if not attendance.office_id:
                attendance.office_id = attendance.user.office_id


        elif attendance.checkout_time is None:
            # Check-out flow
            attendance.checkout_time = datetime.now(tokyo_tz)
            attendance.qr_token = None
            message = "Check-out successful"
            if not attendance.office_id:
                attendance.office_id = attendance.user.office_id


        else:
            # Already checked in and out
            return jsonify({'success': False, 'message': 'Already checked out'}), 400
        
        # Ensure office_id is set based on the user's office at the time
        

        db.session.commit()
        return jsonify({'success': True, 'message': message})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Server error: {str(e)}'}), 500
    


@app.route('/api/generate_qr', methods=['POST'])
def generate_qr():
    user_email = session.get('email')
    user = User.query.filter_by(email=user_email).first()
    
    if not user:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401

    today = date.today()

    # Find latest attendance entry
    latest_attendance = (
        Attendance.query.filter_by(user_id=user.id)
        .order_by(Attendance.id.desc())
        .first()
    )

    # ✅ Create new only if last one is missing or already checked out
    if not latest_attendance or latest_attendance.checkout_time:
        new_attendance = Attendance(
            user_id=user.id,
            date=today,
            qr_token=str(uuid.uuid4()),
            office_id=user.office_id  # ← Add this line to set office at creation
        )

        db.session.add(new_attendance)
        db.session.commit()
        token = new_attendance.qr_token
    else:
        # ❌ Reuse existing (not yet checked out)
        latest_attendance.qr_token = str(uuid.uuid4())
        latest_attendance.office_id=user.office_id
        db.session.commit()
        token = latest_attendance.qr_token

    return jsonify({'success': True, 'token': token})


@app.route('/api/expire_qr', methods=['POST'])
def expire_qr():
    user_email = session.get('email')
    if not user_email:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401

    user = User.query.filter_by(email=user_email).first()
    if not user:
        return jsonify({'success': False, 'message': 'User not found'}), 404

    today = date.today()
    attendance = Attendance.query.filter_by(user_id=user.id, date=today).first()
    if attendance:
        attendance.qr_token = None
        db.session.commit()

    return jsonify({'success': True})


from flask import jsonify, session
from datetime import date
from models import Attendance, User
from flask import jsonify, session
from datetime import date
from models import Attendance, User

from datetime import datetime
import pytz

@app.route('/api/user/attendance-status')
def attendance_status():
    user_email = session.get('email')
    if not user_email:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401

    user = User.query.filter_by(email=user_email).first()
    if not user:
        return jsonify({'success': False, 'message': 'User not found'}), 404

    # Get the most recent attendance record
    record = (Attendance.query
              .filter_by(user_id=user.id)
              .order_by(Attendance.date.desc(), Attendance.id.desc())
              .first())

    # Timezone setup
    jst = pytz.timezone("Asia/Tokyo")

    def to_jst(dt):
        if dt is None:
            return None
        # Ensure UTC awareness
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=pytz.utc)
        return dt.astimezone(jst).isoformat()

    if record:
        checkin_time = to_jst(record.checkin_time)
        checkout_time = to_jst(record.checkout_time)

        if not record.checkin_time or (record.checkin_time and record.checkout_time):
            return jsonify({
                'success': True,
                'checkin_time': checkin_time,
                'checkout_time': checkout_time,
                'enable_checkin': True,
                'enable_checkout': False
            })
        else:
            return jsonify({
                'success': True,
                'checkin_time': checkin_time,
                'checkout_time': None,
                'enable_checkin': False,
                'enable_checkout': True
            })

    return jsonify({
        'success': True,
        'checkin_time': None,
        'checkout_time': None,
        'enable_checkin': True,
        'enable_checkout': False
    })


@app.route('/api/verify_location', methods=['POST'])
def verify_location():
    user_email = session.get('email')
    if not user_email:
        return jsonify({'allowed': False, 'message': 'Unauthorized'}), 401

    user = User.query.filter_by(email=user_email).first()
    if not user or not user.office:
        return jsonify({'allowed': False, 'message': 'No assigned office'}), 403

    data = request.get_json()
    user_lat = data.get('lat')
    user_lng = data.get('lng')

    if user_lat is None or user_lng is None:
        return jsonify({'allowed': False, 'message': 'Missing coordinates'}), 400

    from geopy.distance import geodesic

    office = user.office
    office_coords = (office.latitude, office.longitude)
    user_coords = (user_lat, user_lng)
    print(user_coords)

    if office.latitude is None or office.longitude is None or office.radius_meters is None:
        return jsonify({'allowed': False, 'message': 'Office location is not properly set.'}), 500

    distance = geodesic(user_coords, office_coords).meters
    print(distance)

    if distance <= office.radius_meters:
        return jsonify({'allowed': True})
    else:
        return jsonify({'allowed': False, 'message': f'Outside geofence. Distance: {distance:.2f}m'}), 403

from models import Office
@app.route("/admin_employees", methods=["GET", "POST"])
def admin_employees():
    user_email = session.get('email')
    admin = User.query.filter_by(email=user_email, role="admin").first()
    if not admin:
        return redirect(url_for("index"))

    if request.method == "POST":
        email = request.form.get("email")
        user = User.query.filter_by(email=email).first()

        if user:
            user.manager_id = admin.id
            user.office_id = admin.office_id
            db.session.commit()
        else:
            # Optional: create new placeholder user, or flash error
            flash("User with that email does not exist.", "danger")
            return redirect(url_for("admin_employees"))

        return redirect(url_for("admin_employees"))

    employees = User.query.filter_by(manager_id=admin.id).all()
    return render_template("admin_employees.html", employees=employees)

@app.route("/remove_employee", methods=["POST"])
def remove_employee():
    user_email = session.get('email')
    admin = User.query.filter_by(email=user_email, role="admin").first()
    if not admin:
        return redirect(url_for("index"))

    employee_id = request.form.get("employee_id")
    user = User.query.filter_by(id=employee_id, manager_id=admin.id).first()

    if user:
        db.session.delete(user)
        db.session.commit()

    return redirect(url_for("admin_employees"))


@app.route("/admin_offices", methods=["GET", "POST"])
def admin_offices():  # <-- renamed
    user_email = session.get('email')
    admin = User.query.filter_by(email=user_email, role="admin").first()

    if not admin:
        return redirect(url_for("index"))

    if request.method == "POST":
        name = request.form.get("name")
        lat = request.form.get("latitude")
        lng = request.form.get("longitude")

        if name and lat and lng:
            office = Office(
                name=name,
                latitude=float(lat),
                longitude=float(lng),
                admin_id=admin.id  # initially assigned to creator
            )
            db.session.add(office)
            db.session.commit()
            return redirect(url_for("admin_offices"))

    offices = Office.query.order_by(Office.created_at.desc()).all()
    return render_template("admin_offices.html", offices=offices)


@app.route("/remove_office", methods=["POST"])
def remove_office():
    user_email = session.get('email')
    admin = User.query.filter_by(email=user_email, role="admin").first()
    office_id = request.form.get("office_id")

    office = Office.query.get(office_id)
    if office and office.admin_id == admin.id:
        db.session.delete(office)
        db.session.commit()

    return redirect(url_for("admin_offices"))





@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))


if __name__ == '__main__':
    app.run(port=5000, debug=True)


